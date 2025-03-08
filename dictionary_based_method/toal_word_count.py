import os
import re
import openpyxl
from collections import defaultdict

# Paths
input_folder = "C:/Users/tuant/Downloads/archive/data_txt"  # Folder containing the .txt files
output_file = "total_word_counts1.xlsx"

# Step 1: Initialize total word counts
total_word_counts = defaultdict(int)

# Step 2: Process all .txt files in the specified folder
file_paths = [
    os.path.join(input_folder, file)
    for file in os.listdir(input_folder)
    if file.endswith('.txt')
]

for file_path in file_paths:
    filename = os.path.basename(file_path)
    match = re.match(r"(\d+)_.*_(\d{4})(?:_.*)?\.txt", filename)
    if match:
        file_id, file_year = int(match.group(1)), int(match.group(2))
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read().lower()

            # Count total words in the document
            words_in_file = re.findall(r'\b\w+\b', text)
            total_word_counts[(file_id, file_year)] += len(words_in_file)  # Accumulate counts

# Step 3: Prepare data for Excel
unique_ids = sorted(set(file_id for file_id, _ in total_word_counts.keys()))
unique_years = sorted(set(file_year for _, file_year in total_word_counts.keys()))

# Create a mapping of IDs and years to row and column indices
id_to_row = {file_id: idx + 2 for idx, file_id in enumerate(unique_ids)}
year_to_col = {file_year: idx + 2 for idx, file_year in enumerate(unique_years)}

# Step 4: Write results to an Excel file
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Total Word Counts"

# Write the header row and column
sheet.cell(row=1, column=1, value="File ID/Year")
for col, year in enumerate(unique_years, start=2):
    sheet.cell(row=1, column=col, value=year)
for row, file_id in enumerate(unique_ids, start=2):
    sheet.cell(row=row, column=1, value=file_id)

# Populate the Excel sheet with word counts
for (file_id, file_year), word_count in total_word_counts.items():
    row = id_to_row[file_id]
    col = year_to_col[file_year]
    sheet.cell(row=row, column=col, value=word_count)

# Save the workbook
workbook.save(output_file)
print(f"Total word counts saved to {output_file}")
