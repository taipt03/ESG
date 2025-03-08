import openpyxl

# File paths
file1 = "word_counts_all.xlsx"
file2 = "word_counts_all3.xlsx"
output_file = "word_counts_combined2.xlsx"

# Load both workbooks
wb1 = openpyxl.load_workbook(file1)
wb2 = openpyxl.load_workbook(file2)

# Create a new workbook for the merged data
wb_combined = openpyxl.Workbook()
wb_combined.remove(wb_combined.active)  # Remove the default sheet

# Function to read data from a sheet into a dictionary
def read_sheet_data(sheet):
    data = {}
    years = [sheet.cell(row=1, column=col).value for col in range(2, sheet.max_column + 1)]
    
    for row in range(2, sheet.max_row + 1):
        file_id = sheet.cell(row=row, column=1).value
        if file_id is None:
            continue

        for col, year in enumerate(years, start=2):
            if year is None:
                continue
            value = sheet.cell(row=row, column=col).value
            data.setdefault((file_id, year), value)

    return data

# Merge data from both workbooks
for sheet_name in wb1.sheetnames:
    sheet1 = wb1[sheet_name]
    sheet2 = wb2[sheet_name] if sheet_name in wb2.sheetnames else None

    data1 = read_sheet_data(sheet1)
    data2 = read_sheet_data(sheet2) if sheet2 else {}

    # Merge: Only overwrite if the value in file2 is not "N/A"
    merged_data = data1.copy()
    for key, value in data2.items():
        if value != "N/A":  # Only update if file2 has a real value
            merged_data[key] = value

    # Create a new sheet in the combined workbook
    sheet_combined = wb_combined.create_sheet(title=sheet_name)
    sheet_combined.cell(row=1, column=1, value="ID \\ Year")

    # Extract unique years and IDs
    years = sorted(set(year for _, year in merged_data))
    file_ids = sorted(set(file_id for file_id, _ in merged_data))

    # Write headers (years)
    for col, year in enumerate(years, start=2):
        sheet_combined.cell(row=1, column=col, value=year)

    # Write datax   
    for row, file_id in enumerate(file_ids, start=2):
        sheet_combined.cell(row=row, column=1, value=file_id)
        for col, year in enumerate(years, start=2):
            value = merged_data.get((file_id, year), "N/A")
            sheet_combined.cell(row=row, column=col, value=value)

# Save the merged workbook
wb_combined.save(output_file)
print(f"Merged file saved as {output_file}")
