import os
import re
import time
from collections import defaultdict
import openpyxl
from tqdm import tqdm  # Import tqdm for progress tracking

# Paths
input_folder = "C:/Users/tuant/Downloads/archive/data_txt"  # Folder containing the .txt files
categories_file = "word_categories(ESG).txt"
tone_file = "word_list_old_PN.txt"
output_file = "word_counts_old.xlsx"

# Step 1: Read categories (environmental, governance, social)
with open(categories_file, 'r') as file:
    lines = file.readlines()

categories = {}
for line in lines:
    key, words = line.strip().split(':')
    categories[key.strip()] = {word.strip().lower() for word in words.split(',')}

# Step 2: Read tone categories (positive, negative, forward-looking, risk)
with open(tone_file, 'r') as file:
    lines = file.readlines()

tone_categories = {}
for line in lines:
    key, words = line.strip().split(':')
    tone_categories[key.strip()] = {word.strip().lower() for word in words.split(',')}

# Step 3: Initialize word counts for tone categories
word_counts = {category: defaultdict(lambda: defaultdict(lambda: "N/A")) for category in tone_categories}

# Step 4: Process all .txt files in the specified folder
file_paths = [
    os.path.join(input_folder, file)
    for file in os.listdir(input_folder)
    if file.endswith('.txt')
]

start_time = time.time()  # Start time for tracking duration

for i, file_path in enumerate(tqdm(file_paths, desc="Processing files", unit="file"), start=1):
    filename = os.path.basename(file_path)
    match = re.match(r"(\d+)_.*_(\d{4})(?:_.*)?\.txt", filename)
    
    if match:
        file_id, file_year = int(match.group(1)), int(match.group(2))
        
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read().lower()

            # Tokenize the text into words
            words_in_text = re.findall(r'\b\w+\b', text)

            # Count occurrences of tone words
            for category, tone_words in tone_categories.items():
                count = sum(words_in_text.count(word) for word in tone_words)
                
                if word_counts[category][file_id][file_year] == "N/A":
                    word_counts[category][file_id][file_year] = 0
                word_counts[category][file_id][file_year] += count  # Accumulate counts

    # Estimate remaining time
    elapsed_time = time.time() - start_time
    avg_time_per_file = elapsed_time / i
    remaining_time = avg_time_per_file * (len(file_paths) - i)

# Step 5: Write results to an Excel file
workbook = openpyxl.Workbook()

for tone in tone_categories:
    # Create a sheet for each tone category
    sheet = workbook.create_sheet(title=tone.capitalize())
    sheet.cell(row=1, column=1, value="ID \\ Year")

    # Write the headers dynamically
    files_processed = sorted({(file_id, file_year) for file_id in word_counts[tone] for file_year in word_counts[tone][file_id]})
    file_years = sorted(set(year for _, year in files_processed))
    file_ids = sorted(set(file_id for file_id, _ in files_processed))

    # Populate header row with years
    for col, year in enumerate(file_years, start=2):
        sheet.cell(row=1, column=col, value=year)

    # Write the word counts for each ID and year
    for row, file_id in enumerate(file_ids, start=2):
        sheet.cell(row=row, column=1, value=file_id)
        for col, year in enumerate(file_years, start=2):
            value = word_counts[tone][file_id][year]
            sheet.cell(row=row, column=col, value=value)

# Remove the default sheet if it exists and save the workbook
if "Sheet" in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])
workbook.save(output_file)

print(f"Word counts saved to {output_file}")
