import os
import re
import csv
from collections import defaultdict
import openpyxl
from tqdm import tqdm

# Paths
input_folder = "C:/Users/tuant/Downloads/data/data_txt"  # Folder containing the .txt files
tone_file = "word_list.csv"  # CSV file with 10 columns of words
output_file = "word_counts_tone.xlsx"
# Step 1: Read tone categories from CSV file
tone_categories = defaultdict(set)
with open(tone_file, 'r', encoding='utf-8') as file:
    reader = csv.reader(file)
    headers = next(reader)  # Read the header row for categories
    for row in reader:
        for category, word in zip(headers, row):
            if word.strip():
                tone_categories[category.strip()].add(word.strip().lower())

# Step 2: Initialize word counts for tone categories    
word_counts = {category: defaultdict(lambda: defaultdict(int)) for category in tone_categories}

# Step 3: Process all .txt files in the specified folder
file_paths = [
    os.path.join(input_folder, file)
    for file in os.listdir(input_folder)
    if file.endswith('.txt')
]

for file_path in tqdm(file_paths, desc="Processing files"):
    filename = os.path.basename(file_path)
    match = re.match(r"(\d+)_.*_(\d{4})(?:_.*)?\.txt", filename)
    if match:
        file_id, file_year = int(match.group(1)), int(match.group(2))
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read().lower()

            # Tokenize the text into words
            words_in_text = re.findall(r'\b\w+\b', text)

            # Count occurrences of tone words directly
            for category, tone_words in tone_categories.items():
                word_counts[category][file_id][file_year] = sum(words_in_text.count(word) for word in tone_words)

# Step 4: Write results to an Excel file
workbook = openpyxl.Workbook()

for tone in tqdm(tone_categories, desc="Writing to Excel"):
    # Create a sheet for each tone category
    sheet = workbook.create_sheet(title=tone.capitalize())
    sheet.cell(row=1, column=1, value="ID \\ Year")

    # Write the headers dynamically
    files_processed = sorted({(file_id, file_year) for file_id in word_counts[tone] for file_year in word_counts[tone][file_id]})
    file_years = sorted(set(year for _, year in files_processed))
    file_ids = sorted(set(file_id for file_id, _ in files_processed))

    # Populate header row with years
    for col, year in enumerate(file_years, start=2):
        sheet.cell(row=1, column=col, value=year)

    # Write the word counts for each ID and year
    for row, file_id in enumerate(file_ids, start=2):
        sheet.cell(row=row, column=1, value=file_id)
        for col, year in enumerate(file_years, start=2):
            value = word_counts[tone][file_id][year]
            sheet.cell(row=row, column=col, value=value)

# Remove the default sheet if it exists and save the workbook
if "Sheet" in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])
workbook.save(output_file)

print(f"Word counts saved to {output_file}")


